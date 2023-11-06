import collections
import torch
import torch.nn.functional as F
import utils
import pathlib
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import cv2


class DatasetDistillation:
    def __init__(self, batch, obs_spec, act_spec, discount, nstep, lr_obs, lr_action, device):
        self.batch = batch
        self.obs_spec = obs_spec
        self.act_spec = act_spec
        self.nstep = nstep
        self.lr_obs = lr_obs
        self.lr_action = lr_action
        self.device = device

        discount_ = 1
        for _ in range(nstep):
            discount_ *= discount

        self.init_obs = np.random.rand(1000, *self.obs_spec.shape)
        self.init_action = np.random.rand(1000, *self.act_spec.shape)
        self.init_reward = np.linspace(0, 6, 1001)[:-1].reshape(-1, 1)  # action_repeat * nstep = 6
        self.init_discount = np.full(shape=(1000, 1), fill_value=discount_)
        self.init_next_obs = np.random.rand(1000, *self.obs_spec.shape)

        self.obs = torch.tensor(self.init_obs, dtype=torch.float, requires_grad=True, device=self.device)
        self.action = torch.tensor(self.init_action, dtype=torch.float, requires_grad=True, device=self.device)
        self.reward = torch.tensor(self.init_reward, dtype=torch.float, requires_grad=False, device=self.device)
        self.discount = torch.tensor(self.init_discount, dtype=torch.float, requires_grad=False, device=self.device)
        self.next_obs = torch.tensor(self.init_next_obs, dtype=torch.float, requires_grad=True, device=self.device)

        self.init_obs = self.obs.cpu().detach().numpy()

        self.opt_obs = torch.optim.Adam([self.obs, self.next_obs], lr=self.lr_obs)
        self.opt_action = torch.optim.Adam([self.action], lr=self.lr_action)
        self.index = None
        self.statistic = collections.defaultdict(int)

    def get_data(self, batch, index=True):
        if index:
            reward_ = batch[2].numpy()
            self.index = reward_.reshape(-1)
            for i in range(len(self.index)):
                self.index[i] = np.floor(self.index[i] / 0.006) if self.index[i] != 6.0 else 999  # 6 for reward[5.994]
        obs_ = self.obs[self.index]
        action_ = self.action[self.index]
        reward_ = self.reward[self.index]
        discount_ = self.discount[self.index]
        next_obs_ = self.next_obs[self.index]

        return self.convert(obs_, action_, reward_, discount_, next_obs_)

    def train(self, critic_grad_sync, actor_grad_sync, critic_grad_real, actor_grad_real):
        for i in critic_grad_real:
            i.detach()
        for i in actor_grad_real:
            i.detach()
        metrics = {}
        critic_loss = utils.match_loss(critic_grad_sync, critic_grad_real, self.device, "mse")
        actor_loss = utils.match_loss(actor_grad_sync, actor_grad_real, self.device, "mse")
        loss = critic_loss + actor_loss  # param to actor
        metrics['DD_loss'] = loss.item()
        self.opt_obs.zero_grad(set_to_none=True)
        self.opt_action.zero_grad(set_to_none=True)
        loss.backward()
        # grad = [p.grad for p in [self.obs, self.action, self.reward, self.discount, self.next_obs]]
        self.opt_obs.step()
        self.opt_action.step()

        for i in range(1000):
            self.statistic[f'reward_{i}'] += np.sum(self.index == i)

        return metrics

    def save_img(self, global_step):
        obs_log = self.obs.cpu().detach().numpy()
        img = np.clip(obs_log * 255, 0, 255).astype(np.uint8)
        reward = self.reward.cpu().detach().numpy()
        img_n = np.clip(self.next_obs.cpu().detach().numpy() * 255, 0, 255).astype(np.uint8)
        path = pathlib.Path.cwd() / f'sync_obs/{global_step}'
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        for i in range(1000):
            o = np.transpose(img[i], (1, 2, 0))[:, :, 6:]
            o_n = np.transpose(img_n[i], (1, 2, 0))[:, :, 6:]
            cv2.imwrite(os.path.abspath(path) + f'/img_{i}.png', o)
            cv2.imwrite(os.path.abspath(path) + f'/img_next_{i}.png', o_n)
            # obs = Image(os.path.abspath(path) + f'/img_{i}.png')
            # obs_next = Image(os.path.abspath(path) + f'/img_next_{i}.png')
            # ws.add_image(obs, f'A{i + 1}')
            ws[f'B{i + 1}'] = reward[i][0]
            # ws.add_image(obs_next, f'C{i + 1}')
            ws[f'D{i + 1}'] = self.statistic[f'reward_{i}']
            ws[f'E{i + 1}'] = np.sum(obs_log[i] - self.init_obs[i]) * 255
            ws[f'F{i + 1}'] = np.sum((obs_log[i] - self.init_obs[i]) != 0)
        wb.save(os.path.abspath(path) + f'/sync_data_{global_step}.xlsx')
        # for i in range(1000):
        #     obs = pathlib.Path(os.path.abspath(path) + f'/img_{i}.png')
        #     obs_next = pathlib.Path(os.path.abspath(path) + f'/img_next_{i}.png')
        #     obs.unlink()
        #     obs_next.unlink()

    def convert(self, obs, action, reward, discount, next_obs):
        obs_ = F.normalize(obs, p=2, dim=1)
        action_ = F.normalize(action, p=2, dim=1)
        obs_next_ = F.normalize(next_obs, p=2, dim=1)

        obs_min = torch.tensor(self.obs_spec.minimum, device=self.device)
        obs_max = torch.tensor(self.obs_spec.maximum, device=self.device)
        action_min = torch.tensor(self.act_spec.minimum, device=self.device)
        action_max = torch.tensor(self.act_spec.maximum, device=self.device)
        return (obs_ * (obs_max - obs_min) + obs_min,
                action_ * (action_max - action_min) + action_min, reward, discount,
                obs_next_ * (obs_max - obs_min) + obs_min)
